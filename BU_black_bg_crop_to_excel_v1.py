import io
import re
import shutil
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# 처리 대상 이미지 확장자
ALLOWED_EXTENSIONS = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp")
# 파일명에서 LotID/종류(BU, WU)를 뽑기 위한 패턴
LOT_PATTERN = re.compile(r"^(?P<lotid>.+)_(?P<kind>BU|WU)_\d+$", re.IGNORECASE)


def get_resized_xl_image(image_path: Path, max_width_px: int) -> XLImage | None:
    # 엑셀 파일 용량 다이어트를 위해 삽입 전에 이미지를 리사이징하여 BytesIO로 반환
    if not image_path.exists():
        return None
    try:
        with Image.open(image_path) as img:
            w, h = img.size
            if w > max_width_px:
                ratio = max_width_px / w
                new_w, new_h = int(w * ratio), int(h * ratio)
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format="PNG")
            img_byte_arr.seek(0)
            return XLImage(img_byte_arr)
    except Exception as e:
        print(f"Error resizing image {image_path}: {e}")
        return None


def print_progress(label: str, current: int, total: int, done: bool = False) -> None:
    # 진행률 표시 공통 함수
    if total <= 0:
        return
    percent = (current / total) * 100
    end = "\n" if done else "\r"
    print(f"{label}: {current}/{total} ({percent:5.1f}%)", end=end, flush=True)


def unique_file_path(path: Path) -> Path:
    # 동일 파일명이 이미 있으면 파일명 뒤에 _1, _2 ... 를 붙여 저장 경로를 만든다.
    if not path.exists():
        return path

    parent = path.parent
    stem = path.stem
    suffix = path.suffix
    idx = 1
    while True:
        candidate = parent / f"{stem}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def ask_int(prompt: str, default: int) -> int:
    # 숫자 입력(엔터면 기본값 사용)
    raw = input(f"{prompt} (기본값 {default}): ").strip().replace('"', "")
    if not raw:
        return default
    return int(raw)


def find_non_black_bbox(img: Image.Image, threshold: int = 12):
    # 검은 배경(저밝기)을 제외한 영역의 최소 사각형(BBox) 검출 (NumPy 최적화 버전)
    arr = np.array(img.convert("L"))
    rows = np.any(arr > threshold, axis=1)
    cols = np.any(arr > threshold, axis=0)

    if not np.any(rows) or not np.any(cols):
        return None

    min_y, max_y = np.where(rows)[0][[0, -1]]
    min_x, max_x = np.where(cols)[0][[0, -1]]

    return int(min_x), int(min_y), int(max_x + 1), int(max_y + 1)


def add_padding(box, w: int, h: int, pad: int):
    # 잘림 방지를 위해 크롭 박스에 여백(padding) 추가
    left, top, right, bottom = box
    return (
        max(0, left - pad),
        max(0, top - pad),
        min(w, right + pad),
        min(h, bottom + pad),
    )


def parse_lot_kind(stem: str):
    # 파일명에서 LotID와 BU/WU를 파싱
    # 패턴 불일치 시 kind=UNKNOWN으로 처리
    m = LOT_PATTERN.match(stem)
    if not m:
        return stem, "UNKNOWN"
    return m.group("lotid"), m.group("kind").upper()


def crop_images(input_root: Path, output_root: Path, threshold: int, padding: int):
    # 입력 폴더 전체 이미지를 자동 크롭하여 output_root에 저장
    if output_root.exists():
        shutil.rmtree(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    image_files = [
        p for p in input_root.rglob("*") if p.is_file() and p.suffix.lower() in ALLOWED_EXTENSIONS
    ]
    total_images = len(image_files)
    print(f"\n[1/2] 이미지 크롭 시작 (대상: {total_images}개)")

    records = []
    for idx, src in enumerate(image_files, start=1):
        rel = src.relative_to(input_root)
        dst_raw = output_root / rel
        dst = dst_raw
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst = unique_file_path(dst)
        renamed = dst != dst_raw

        lot_id, kind = parse_lot_kind(src.stem)

        try:
            with Image.open(src) as img:
                bbox = find_non_black_bbox(img, threshold=threshold)
                if bbox is None:
                    cropped = img.copy()
                    used_bbox = (0, 0, img.width, img.height)
                    status = "NO_OBJECT_DETECTED"
                else:
                    used_bbox = add_padding(bbox, img.width, img.height, padding)
                    cropped = img.crop(used_bbox)
                    status = "OK"

                if dst.suffix.lower() in (".jpg", ".jpeg") and cropped.mode not in ("RGB", "L"):
                    cropped = cropped.convert("RGB")
                cropped.save(dst)

            records.append(
                {
                    "lot_id": lot_id,
                    "kind": kind,
                    "src": src,
                    "dst": dst,
                    "bbox": used_bbox,
                    "status": status,
                    "renamed_on_save": "TRUE" if renamed else "FALSE",
                }
            )
        except Exception as exc:
            records.append(
                {
                    "lot_id": lot_id,
                    "kind": kind,
                    "src": src,
                    "dst": None,
                    "bbox": None,
                    "status": f"ERROR: {exc}",
                    "renamed_on_save": "FALSE",
                }
            )

        if idx == 1 or idx % 10 == 0 or idx == total_images:
            print_progress("  크롭 진행", idx, total_images, done=(idx == total_images))

    return records


def write_excel(records, excel_path: Path, image_width_px: int = 240):
    # 결과 시트: LotID별 BU/WU 이미지 배치 (사용자 입력 컬럼은 공란 유지)
    # 상세 시트: 경로/중복 정보 정리 (이미지 없음)
    wb = Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(["LotID", "판정", "BU data", "BU Image", "WU data", "WU Image"])

    total = len(records)
    print(f"\n[2/2] 엑셀 작성 시작 (행: {total}개)")

    grouped: dict[str, dict[str, Path | None]] = {}
    for rec in records:
        lot_id = rec["lot_id"]
        kind = rec["kind"]
        grouped.setdefault(lot_id, {"BU": None, "WU": None})
        if kind in ("BU", "WU") and rec["dst"] is not None and grouped[lot_id][kind] is None:
            grouped[lot_id][kind] = rec["dst"]

    row = 2
    for lot_id in sorted(grouped.keys()):
        ws.cell(row=row, column=1, value=lot_id)
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=5, value="")

        bu_path = grouped[lot_id]["BU"]
        wu_path = grouped[lot_id]["WU"]
        max_img_height = 0

        if bu_path is not None and Path(bu_path).exists():
            bu_img = get_resized_xl_image(Path(bu_path), image_width_px)
            if bu_img:
                ws.add_image(bu_img, f"D{row}")
                max_img_height = max(max_img_height, bu_img.height)

        if wu_path is not None and Path(wu_path).exists():
            wu_img = get_resized_xl_image(Path(wu_path), image_width_px)
            if wu_img:
                ws.add_image(wu_img, f"F{row}")
                max_img_height = max(max_img_height, wu_img.height)

        ws.row_dimensions[row].height = max(25, int(max_img_height * 0.75))
        row += 1

    detail_ws = wb.create_sheet("경로_중복정리")
    detail_ws.append(
        ["LotID", "Kind", "Status", "RenamedOnSave", "CropBBox", "SourcePath", "CroppedPath"]
    )

    for idx, rec in enumerate(records, start=1):
        bbox_text = "" if rec["bbox"] is None else str(rec["bbox"])
        dst_text = "" if rec["dst"] is None else str(rec["dst"])
        detail_ws.append(
            [
                rec["lot_id"],
                rec["kind"],
                rec["status"],
                rec.get("renamed_on_save", "FALSE"),
                bbox_text,
                str(rec["src"]),
                dst_text,
            ]
        )
        if idx == 1 or idx % 10 == 0 or idx == total:
            print_progress("  엑셀 진행", idx, total, done=(idx == total))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 36
    detail_ws.column_dimensions["A"].width = 26
    detail_ws.column_dimensions["B"].width = 10
    detail_ws.column_dimensions["C"].width = 24
    detail_ws.column_dimensions["D"].width = 16
    detail_ws.column_dimensions["E"].width = 24
    detail_ws.column_dimensions["F"].width = 60
    detail_ws.column_dimensions["G"].width = 60
    wb.save(excel_path)


def main():
    # 실행 순서:
    # 1) 입력 폴더/파라미터 수집
    # 2) 자동 크롭
    # 3) 엑셀 리포트 생성
    print("\n--- 검은 배경 이미지 자동 크롭 + 엑셀 삽입 v1 ---")
    raw = input("👉 입력 폴더 경로: ").strip().replace('"', "")
    input_root = Path(raw)
    if not input_root.exists() or not input_root.is_dir():
        print(f"🚨 폴더가 없어: {input_root}")
        return

    threshold = ask_int("비검정 판정 임계값(0~255)", 12)
    padding = ask_int("크롭 패딩(px)", 8)

    output_root = input_root.parent / f"{input_root.name}_cropped_v1"
    excel_path = output_root / "crop_report.xlsx"

    print(f"\n📌 입력: {input_root}")
    print(f"📌 출력: {output_root}")
    print(f"📌 임계값: {threshold}, 패딩: {padding}")

    records = crop_images(input_root, output_root, threshold, padding)
    write_excel(records, excel_path)

    ok_count = sum(1 for r in records if r["status"] == "OK")
    nodetect_count = sum(1 for r in records if r["status"] == "NO_OBJECT_DETECTED")
    error_count = sum(1 for r in records if r["status"].startswith("ERROR"))

    print("\n--- 결과 ---")
    print(f"전체 이미지: {len(records)}")
    print(f"크롭 성공: {ok_count}")
    print(f"객체 미검출(원본 유지): {nodetect_count}")
    print(f"오류: {error_count}")
    print(f"엑셀: {excel_path}")
    print("완료!")


if __name__ == "__main__":
    main()
