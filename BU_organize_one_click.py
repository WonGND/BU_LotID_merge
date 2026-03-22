import csv
import io
import re
import shutil
from datetime import datetime
from pathlib import Path

import cv2
import matplotlib
import numpy as np
import pandas as pd
matplotlib.use("Agg")
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from matplotlib import pyplot as plt
from matplotlib.patches import Rectangle
from PIL import Image, ImageDraw

# 처리 대상 이미지 확장자
ALLOWED_EXTENSIONS = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp")
# 파일명에서 LotID/종류(BU, WU)를 뽑기 위한 패턴
LOT_PATTERN = re.compile(r"^(?P<lotid>.+)_(?P<kind>BU|WU)_\d+$", re.IGNORECASE)
DATA_FILE_PATTERN = "LMK6DataLog.csv"
BU_SPEC_MIN = 50.0
WU_SPEC_MIN = 80.0
BU_GRID_COLS = 48
BU_GRID_ROWS = 27
DETAIL_ROW_HEIGHT = 22
INNER_TRIM_VARIANTS = (5,)
WORST_POINT_EDGE_MARGIN_CELLS_X = 2
WORST_POINT_EDGE_MARGIN_CELLS_Y = 2
PRODUCT_CELL_CONTENT_RATIO_MIN = 0.7
PRODUCT_ROW_COVERAGE_MIN = 0.08
PRODUCT_COL_COVERAGE_MIN = 0.08
PRODUCT_BOUND_EXPAND_CELLS = 1
GRID_REFINED_MARGIN_PX = 10
MODEL_NAME_CANDIDATES = (
    "Model_Name",
    "ModelName",
    "Model",
    "RecipeName",
    "Recipe",
    "Product",
    "Product_Name",
)


class PipelineCancelled(Exception):
    pass


def get_resized_xl_image(image_path: Path, max_width_px: int) -> XLImage | None:
    # 엑셀 파일 용량 다이어트를 위해 삽입 전에 이미지를 리사이징하여 BytesIO로 반환
    if not image_path.exists():
        return None
    try:
        with Image.open(image_path) as img:
            w, h = img.size
            if w > max_width_px:
                ratio = max_width_px / w
                new_w, new_h = int(w * ratio), int(h * ratio)
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            
            img_byte_arr = io.BytesIO()
            # 이미지 형식을 유지하되, 리사이징 시에는 최적화된 포맷으로 저장
            img.save(img_byte_arr, format="PNG")
            img_byte_arr.seek(0)
            return XLImage(img_byte_arr)
    except Exception as e:
        print(f"Error resizing image {image_path}: {e}")
        return None


def print_progress(label: str, current: int, total: int, done: bool = False) -> None:
    # 진행률 표시 공통 함수
    if total <= 0:
        return
    percent = (current / total) * 100
    print(f"{label}: {current}/{total} ({percent:5.1f}%)", flush=True)


def ensure_not_cancelled(cancel_check=None) -> None:
    if cancel_check and cancel_check():
        raise PipelineCancelled("사용자 요청으로 작업이 중지되었습니다.")


def unique_folder_path(base_dir: Path, folder_name: str) -> Path:
    # 동일 폴더명이 이미 있으면 _1, _2 ... 를 붙여 새 경로를 만든다.
    candidate = base_dir / folder_name
    if not candidate.exists():
        return candidate

    idx = 1
    while True:
        candidate = base_dir / f"{folder_name}_{idx}"
        if not candidate.exists():
            return candidate
        idx += 1


def unique_file_path(path: Path) -> Path:
    # 동일 파일명이 이미 있으면 파일명 뒤에 _1, _2 ... 를 붙여 저장 경로를 만든다.
    if not path.exists():
        return path

    parent = path.parent
    stem = path.stem
    suffix = path.suffix
    idx = 1
    while True:
        candidate = parent / f"{stem}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def ask_int(prompt: str, default: int) -> int:
    # 숫자 입력(엔터면 기본값 사용)
    raw = input(f"{prompt} (기본값 {default}): ").strip().replace('"', "")
    if not raw:
        return default
    return int(raw)


def ask_path(prompt: str) -> Path:
    return Path(input(prompt).strip().replace('"', ""))


def folder_time_key(path: Path) -> tuple[float, float]:
    # 최신 폴더 비교 기준: 생성시각 우선, 동일하면 수정시각
    stat = path.stat()
    return (stat.st_ctime, stat.st_mtime)


def is_lotid_folder(path: Path) -> bool:
    # LotID 폴더 판정 규칙: 이미지 파일이 1개 이상 있는 디렉터리
    if not path.is_dir():
        return False
    image_count = 0
    for child in path.iterdir():
        if child.is_file() and child.suffix.lower() in ALLOWED_EXTENSIONS:
            image_count += 1
    return image_count >= 1


def format_ts(ts: float) -> str:
    # CSV 가독성을 위해 타임스탬프를 날짜 문자열로 변환
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def parse_lmk_time(value: str) -> datetime:
    # LMK 로그 시간 형식: 2026.02.24 09:30:51
    return datetime.strptime(value.strip(), "%Y.%m.%d %H:%M:%S")


def format_measurement_value(value: str) -> str:
    # 엑셀에는 핵심 수치만 간단히 넣는다.
    if value is None:
        return ""
    return str(value).strip()


def extract_model_name(row: dict) -> str:
    for key in MODEL_NAME_CANDIDATES:
        value = (row.get(key) or "").strip()
        if value:
            return value
    return ""


def to_float(value) -> float | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def iter_csv_dict_rows(csv_path: Path):
    # CSV 인코딩이 파일마다 다를 수 있어서 순차적으로 시도한다.
    encodings = ("utf-8-sig", "cp949", "euc-kr", "utf-8")
    last_error = None

    for encoding in encodings:
        try:
            with csv_path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            return rows
        except UnicodeDecodeError as exc:
            last_error = exc

    raise UnicodeDecodeError(
        getattr(last_error, "encoding", "unknown"),
        getattr(last_error, "object", b""),
        getattr(last_error, "start", 0),
        getattr(last_error, "end", 1),
        f"{csv_path} 파일 인코딩을 읽지 못함",
    )


def collect_latest_measurements(data_root: Path, cancel_check=None) -> tuple[dict[str, dict], list[dict]]:
    # 여러 LMK6DataLog.csv를 재귀 탐색해 Panel_ID 기준 최신 행만 남긴다.
    csv_files = sorted(data_root.rglob(DATA_FILE_PATTERN))
    total_files = len(csv_files)
    print(f"\n[4/7] 측정 CSV 스캔 시작 (대상 파일: {total_files}개)")

    latest_by_lotid: dict[str, dict] = {}
    rows_for_detail: list[dict] = []
    if total_files == 0:
        print("  측정 CSV를 찾지 못해서 데이터 입력은 공란으로 둔다.")
        return latest_by_lotid, rows_for_detail

    for file_idx, csv_path in enumerate(csv_files, start=1):
        ensure_not_cancelled(cancel_check)
        if file_idx == 1 or file_idx % 10 == 0 or file_idx == total_files:
            print_progress("  CSV 스캔 진행", file_idx, total_files, done=(file_idx == total_files))

        for row in iter_csv_dict_rows(csv_path):
            ensure_not_cancelled(cancel_check)
            lot_id = (row.get("Panel_ID") or "").strip()
            if not lot_id:
                continue

            measured_at_raw = row.get("Time", "")
            measured_at = parse_lmk_time(measured_at_raw)
            record = {
                "lot_id": lot_id,
                "model_name": extract_model_name(row),
                "judge": (row.get("Judge") or "").strip(),
                "black_uniformity": format_measurement_value(row.get("Black_Uniformity")),
                "white_uniformity": format_measurement_value(row.get("White_Uniformity")),
                "time_raw": measured_at_raw,
                "time_obj": measured_at,
                "source_file": str(csv_path),
            }

            current = latest_by_lotid.get(lot_id)
            is_latest = current is None or measured_at >= current["time_obj"]
            if is_latest:
                latest_by_lotid[lot_id] = record

            rows_for_detail.append(
                {
                    "lot_id": lot_id,
                    "model_name": record["model_name"],
                    "judge": record["judge"],
                    "black_uniformity": record["black_uniformity"],
                    "white_uniformity": record["white_uniformity"],
                    "time_raw": measured_at_raw,
                    "source_file": str(csv_path),
                    "selected_latest_final": "FALSE",
                }
            )

    latest_signatures = {
        (value["lot_id"], value["time_raw"], value["source_file"]) for value in latest_by_lotid.values()
    }
    for row in rows_for_detail:
        row["selected_latest_final"] = (
            "TRUE"
            if (row["lot_id"], row["time_raw"], row["source_file"]) in latest_signatures
            else "FALSE"
        )

    return latest_by_lotid, rows_for_detail


def build_metric_summary(latest_measurements: dict[str, dict], key: str, spec_min: float) -> dict:
    values = [to_float(row.get(key)) for row in latest_measurements.values()]
    values = [v for v in values if v is not None]
    if not values:
        return {
            "count": 0,
            "pass_count": 0,
            "fail_count": 0,
            "min": None,
            "avg": None,
            "max": None,
            "median": None,
            "sorted_values": [],
        }

    sorted_values = sorted(values)
    count = len(sorted_values)
    mid = count // 2
    if count % 2 == 0:
        median = (sorted_values[mid - 1] + sorted_values[mid]) / 2
    else:
        median = sorted_values[mid]

    pass_count = sum(1 for value in sorted_values if value >= spec_min)
    return {
        "count": count,
        "pass_count": pass_count,
        "fail_count": count - pass_count,
        "min": min(sorted_values),
        "avg": sum(sorted_values) / count,
        "max": max(sorted_values),
        "median": median,
        "sorted_values": sorted_values,
    }


def build_distribution(values: list[float], bins: list[tuple[str, float, float]]) -> list[tuple[str, int]]:
    counts: list[tuple[str, int]] = []
    for label, lower, upper in bins:
        count = sum(1 for value in values if lower <= value < upper)
        counts.append((label, count))
    return counts


def write_card(ws, top_left: str, title: str, value, subtitle: str, fill_color: str) -> None:
    start_col = ws[top_left].column
    start_row = ws[top_left].row
    end_col = start_col + 2
    end_row = start_row + 2
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = f"{title}\n{value}\n{subtitle}"
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_kpi_table(ws, start_row: int, start_col: int, title: str, rows: list[tuple[str, object]]) -> None:
    ws.cell(row=start_row, column=start_col, value=title)
    ws.cell(row=start_row, column=start_col).font = Font(size=12, bold=True, color="1F2937")
    for offset, (label, value) in enumerate(rows, start=1):
        ws.cell(row=start_row + offset, column=start_col, value=label)
        ws.cell(row=start_row + offset, column=start_col + 1, value=value)
        ws.cell(row=start_row + offset, column=start_col).fill = PatternFill("solid", fgColor="F3F4F6")
        ws.cell(row=start_row + offset, column=start_col).font = Font(bold=True, color="374151")


def style_line_chart(chart: LineChart, value_color: str, spec_color: str) -> None:
    chart.style = 10
    chart.legend.position = "b"
    chart.height = 7.8
    chart.width = 13.5
    chart.smooth = True
    if len(chart.ser) >= 1:
        chart.ser[0].graphicalProperties.line.solidFill = value_color
        chart.ser[0].graphicalProperties.line.width = 22000
        chart.ser[0].marker.symbol = "circle"
        chart.ser[0].marker.size = 6
    if len(chart.ser) >= 2:
        chart.ser[1].graphicalProperties.line.solidFill = spec_color
        chart.ser[1].graphicalProperties.line.prstDash = "dash"
        chart.ser[1].graphicalProperties.line.width = 14000


def style_bar_chart(chart: BarChart, fill_color: str) -> None:
    chart.style = 11
    chart.legend = None
    chart.height = 7.6
    chart.width = 11.8
    if len(chart.ser) >= 1:
        chart.ser[0].graphicalProperties.solidFill = fill_color
        chart.ser[0].graphicalProperties.line.solidFill = fill_color


def style_pie_chart(chart: PieChart) -> None:
    chart.style = 26
    chart.legend.position = "b"
    chart.height = 6.4
    chart.width = 8.4


def pick_worst_lotids(latest_measurements: dict[str, dict], key: str, limit: int = 10) -> list[tuple[str, str, float]]:
    rows = []
    for lot_id, measurement in latest_measurements.items():
        value = to_float(measurement.get(key))
        if value is None:
            continue
        rows.append((lot_id, measurement.get("judge", ""), value))
    rows.sort(key=lambda item: item[2])
    return rows[:limit]


def add_visualization_sheet(wb: Workbook, latest_measurements: dict[str, dict]) -> None:
    ws = wb.create_sheet("시각화")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "BU / WU Measurement Dashboard"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["I1"] = datetime.now().strftime("Updated: %Y-%m-%d %H:%M:%S")
    ws["I1"].font = Font(size=10, color="6B7280")

    bu_summary = build_metric_summary(latest_measurements, "black_uniformity", BU_SPEC_MIN)
    wu_summary = build_metric_summary(latest_measurements, "white_uniformity", WU_SPEC_MIN)
    total_count = max(bu_summary["count"], wu_summary["count"])

    write_card(ws, "A4", "전체 LotID", total_count, "최신 측정 기준", "0F766E")
    write_card(ws, "D4", "BU Fail", bu_summary["fail_count"], f"Spec {BU_SPEC_MIN}", "B91C1C")
    write_card(ws, "G4", "WU Fail", wu_summary["fail_count"], f"Spec {WU_SPEC_MIN}", "7C2D12")
    write_card(ws, "J4", "Pass Rate", f"{(0 if total_count == 0 else ((bu_summary['pass_count'] + wu_summary['pass_count']) / max(1, bu_summary['count'] + wu_summary['count']) * 100)):.1f}%", "BU+WU combined", "1D4ED8")

    ws["A8"] = "Metric Summary"
    ws["A8"].font = Font(size=13, bold=True, color="111827")
    ws["A9"] = "지표"
    ws["B9"] = "Spec Min"
    ws["C9"] = "Count"
    ws["D9"] = "Pass"
    ws["E9"] = "Fail"
    ws["F9"] = "Min"
    ws["G9"] = "Avg"
    ws["H9"] = "Median"
    ws["I9"] = "Max"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws["9:9"]:
        cell.fill = header_fill
        cell.font = header_font

    metric_rows = [
        ("Black Uniformity", BU_SPEC_MIN, bu_summary),
        ("White Uniformity", WU_SPEC_MIN, wu_summary),
    ]
    for row_idx, (label, spec_min, summary) in enumerate(metric_rows, start=10):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=spec_min)
        ws.cell(row=row_idx, column=3, value=summary["count"])
        ws.cell(row=row_idx, column=4, value=summary["pass_count"])
        ws.cell(row=row_idx, column=5, value=summary["fail_count"])
        ws.cell(row=row_idx, column=6, value=summary["min"])
        ws.cell(row=row_idx, column=7, value=summary["avg"])
        ws.cell(row=row_idx, column=8, value=summary["median"])
        ws.cell(row=row_idx, column=9, value=summary["max"])

    # 정렬 추세 차트용 데이터
    ws["K9"] = "BU_index"
    ws["L9"] = "BU_value"
    ws["M9"] = "BU_spec"
    for idx, value in enumerate(bu_summary["sorted_values"], start=10):
        ws.cell(row=idx, column=11, value=idx - 2)
        ws.cell(row=idx, column=12, value=value)
        ws.cell(row=idx, column=13, value=BU_SPEC_MIN)

    ws["O9"] = "WU_index"
    ws["P9"] = "WU_value"
    ws["Q9"] = "WU_spec"
    for idx, value in enumerate(wu_summary["sorted_values"], start=10):
        ws.cell(row=idx, column=15, value=idx - 2)
        ws.cell(row=idx, column=16, value=value)
        ws.cell(row=idx, column=17, value=WU_SPEC_MIN)

    bu_line = LineChart()
    bu_line.title = "BU 분포 추세"
    bu_line.y_axis.title = "Black Uniformity"
    bu_line.x_axis.title = "정렬 순서"
    bu_data = Reference(ws, min_col=12, max_col=13, min_row=9, max_row=max(10, len(bu_summary["sorted_values"]) + 9))
    bu_cats = Reference(ws, min_col=11, min_row=10, max_row=max(10, len(bu_summary["sorted_values"]) + 9))
    bu_line.add_data(bu_data, titles_from_data=True)
    bu_line.set_categories(bu_cats)
    style_line_chart(bu_line, "DC2626", "94A3B8")
    ws.add_chart(bu_line, "A14")

    wu_line = LineChart()
    wu_line.title = "WU 분포 추세"
    wu_line.y_axis.title = "White Uniformity"
    wu_line.x_axis.title = "정렬 순서"
    wu_data = Reference(ws, min_col=16, max_col=17, min_row=9, max_row=max(10, len(wu_summary["sorted_values"]) + 9))
    wu_cats = Reference(ws, min_col=15, min_row=10, max_row=max(10, len(wu_summary["sorted_values"]) + 9))
    wu_line.add_data(wu_data, titles_from_data=True)
    wu_line.set_categories(wu_cats)
    style_line_chart(wu_line, "16A34A", "94A3B8")
    ws.add_chart(wu_line, "N14")

    # Pass/Fail 파이 차트
    ws["A31"] = "Metric"
    ws["B31"] = "Pass"
    ws["C31"] = "Fail"
    ws["A32"] = "BU"
    ws["B32"] = bu_summary["pass_count"]
    ws["C32"] = bu_summary["fail_count"]
    ws["A33"] = "WU"
    ws["B33"] = wu_summary["pass_count"]
    ws["C33"] = wu_summary["fail_count"]

    bu_pie = PieChart()
    bu_pie.title = "BU Pass/Fail"
    bu_pie.add_data(Reference(ws, min_col=2, max_col=3, min_row=32, max_row=32), from_rows=True)
    bu_pie.set_categories(Reference(ws, min_col=2, max_col=3, min_row=31, max_row=31))
    bu_pie.dataLabels = DataLabelList()
    bu_pie.dataLabels.showPercent = True
    bu_pie.dataLabels.showVal = True
    style_pie_chart(bu_pie)
    ws.add_chart(bu_pie, "A35")

    wu_pie = PieChart()
    wu_pie.title = "WU Pass/Fail"
    wu_pie.add_data(Reference(ws, min_col=2, max_col=3, min_row=33, max_row=33), from_rows=True)
    wu_pie.set_categories(Reference(ws, min_col=2, max_col=3, min_row=31, max_row=31))
    wu_pie.dataLabels = DataLabelList()
    wu_pie.dataLabels.showPercent = True
    wu_pie.dataLabels.showVal = True
    style_pie_chart(wu_pie)
    ws.add_chart(wu_pie, "J35")

    # Spec 기준 중심 버킷 분포
    bu_bins = [
        ("<40", 0, 40),
        ("40-45", 40, 45),
        ("45-50", 45, 50),
        ("50-55", 50, 55),
        ("55-60", 55, 60),
        ("60+", 60, 9999),
    ]
    wu_bins = [
        ("<70", 0, 70),
        ("70-75", 70, 75),
        ("75-80", 75, 80),
        ("80-85", 80, 85),
        ("85-90", 85, 90),
        ("90+", 90, 9999),
    ]
    bu_distribution = build_distribution(bu_summary["sorted_values"], bu_bins)
    wu_distribution = build_distribution(wu_summary["sorted_values"], wu_bins)

    ws["S9"] = "BU_bucket"
    ws["T9"] = "BU_count"
    for idx, (label, count) in enumerate(bu_distribution, start=10):
        ws.cell(row=idx, column=19, value=label)
        ws.cell(row=idx, column=20, value=count)

    ws["V9"] = "WU_bucket"
    ws["W9"] = "WU_count"
    for idx, (label, count) in enumerate(wu_distribution, start=10):
        ws.cell(row=idx, column=22, value=label)
        ws.cell(row=idx, column=23, value=count)

    bu_bar = BarChart()
    bu_bar.title = "BU 분포 구간"
    bu_bar.y_axis.title = "Count"
    bu_bar.x_axis.title = "Range"
    bu_bar.add_data(Reference(ws, min_col=20, min_row=9, max_row=9 + len(bu_distribution)), titles_from_data=True)
    bu_bar.set_categories(Reference(ws, min_col=19, min_row=10, max_row=9 + len(bu_distribution)))
    style_bar_chart(bu_bar, "DC2626")
    ws.add_chart(bu_bar, "T14")

    wu_bar = BarChart()
    wu_bar.title = "WU 분포 구간"
    wu_bar.y_axis.title = "Count"
    wu_bar.x_axis.title = "Range"
    wu_bar.add_data(Reference(ws, min_col=23, min_row=9, max_row=9 + len(wu_distribution)), titles_from_data=True)
    wu_bar.set_categories(Reference(ws, min_col=22, min_row=10, max_row=9 + len(wu_distribution)))
    style_bar_chart(wu_bar, "16A34A")
    ws.add_chart(wu_bar, "T35")

    write_kpi_table(
        ws,
        31,
        20,
        "Worst BU LotID",
        [(lot_id, value) for lot_id, judge, value in pick_worst_lotids(latest_measurements, "black_uniformity")],
    )
    write_kpi_table(
        ws,
        31,
        23,
        "Worst WU LotID",
        [(lot_id, value) for lot_id, judge, value in pick_worst_lotids(latest_measurements, "white_uniformity")],
    )

    ws.column_dimensions["A"].width = 22
    for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "S", "T", "V", "W"):
        ws.column_dimensions[col].width = 12


def collect_latest_lotid_folders(integrated_root: Path, cancel_check=None) -> tuple[dict[str, Path], list[dict]]:
    # 전체 폴더를 훑어서 LotID별 최신 폴더 1개만 남긴다.
    latest_by_lotid: dict[str, Path] = {}
    rows: list[dict] = []

    dir_candidates = [p for p in integrated_root.rglob("*") if p.is_dir()]
    total_dirs = len(dir_candidates)
    print(f"\n[1/7] LotID 폴더 스캔 시작 (대상 폴더: {total_dirs}개)")

    for idx, p in enumerate(dir_candidates, start=1):
        ensure_not_cancelled(cancel_check)
        if idx == 1 or idx % 50 == 0 or idx == total_dirs:
            print_progress("  스캔 진행", idx, total_dirs, done=(idx == total_dirs))

        if not is_lotid_folder(p):
            continue

        lot_id = p.name
        current = latest_by_lotid.get(lot_id)
        is_latest = False

        if current is None or folder_time_key(p) > folder_time_key(current):
            latest_by_lotid[lot_id] = p
            is_latest = True

        created_ts = p.stat().st_ctime
        modified_ts = p.stat().st_mtime
        rows.append(
            {
                "lot_id": lot_id,
                "folder_path": str(p),
                "created_time": format_ts(created_ts),
                "modified_time": format_ts(modified_ts),
                "selected_latest_at_scan_time": "TRUE" if is_latest else "FALSE",
            }
        )

    selected_paths = {str(v) for v in latest_by_lotid.values()}
    for row in rows:
        row["selected_latest_final"] = "TRUE" if row["folder_path"] in selected_paths else "FALSE"

    return latest_by_lotid, rows


def copy_latest_folders(latest_by_lotid: dict[str, Path], output_root: Path, cancel_check=None) -> None:
    # 최종 선택된 LotID 폴더만 결과 폴더로 복사
    if output_root.exists():
        shutil.rmtree(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    items = sorted(latest_by_lotid.items())
    total = len(items)
    print(f"\n[2/7] 최신 LotID 폴더 복사 시작 (대상: {total}개)")

    for idx, (lot_id, src) in enumerate(items, start=1):
        ensure_not_cancelled(cancel_check)
        dst = unique_folder_path(output_root, lot_id)
        shutil.copytree(src, dst)
        if idx == 1 or idx % 20 == 0 or idx == total:
            print_progress("  복사 진행", idx, total, done=(idx == total))


def write_merge_report(rows: list[dict], output_root: Path) -> Path:
    # 병합(merge) 판단 결과를 CSV로 저장
    report_path = output_root / "merge_report.csv"
    fieldnames = [
        "lot_id",
        "folder_path",
        "created_time",
        "modified_time",
        "selected_latest_at_scan_time",
        "selected_latest_final",
    ]
    print("\n[3/7] Merge 리포트 저장")
    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print("  Merge 리포트 저장 완료")
    return report_path


def find_non_black_bbox(img: Image.Image, threshold: int = 12):
    # 검은 배경(저밝기)을 제외한 영역의 최소 사각형(BBox) 검출 (NumPy 최적화 버전)
    arr = np.array(img.convert("L"))
    rows = np.any(arr > threshold, axis=1)
    cols = np.any(arr > threshold, axis=0)

    if not np.any(rows) or not np.any(cols):
        return None

    min_y, max_y = np.where(rows)[0][[0, -1]]
    min_x, max_x = np.where(cols)[0][[0, -1]]

    return int(min_x), int(min_y), int(max_x + 1), int(max_y + 1)


def load_cv2_image(image_path: Path):
    data = np.fromfile(str(image_path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def get_refined_product_bbox(image_path: Path, margin_px: int = 10):
    img = load_cv2_image(image_path)
    if img is None:
        return None

    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower_bound = np.array([0, 50, 50], dtype=np.uint8)
    upper_bound = np.array([180, 255, 255], dtype=np.uint8)
    mask = cv2.inRange(hsv, lower_bound, upper_bound)

    kernel = np.ones((5, 5), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None

    max_cnt = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(max_cnt)
    x_new = max(0, x + margin_px)
    y_new = max(0, y + margin_px)
    w_new = max(1, w - (margin_px * 2))
    h_new = max(1, h - (margin_px * 2))

    return (
        x_new,
        y_new,
        min(img.shape[1], x_new + w_new),
        min(img.shape[0], y_new + h_new),
    )


def add_padding(box, w: int, h: int, pad: int):
    # 잘림 방지를 위해 크롭 박스에 여백(padding) 추가
    left, top, right, bottom = box
    return (
        max(0, left - pad),
        max(0, top - pad),
        min(w, right + pad),
        min(h, bottom + pad),
    )


def parse_lot_kind(stem: str):
    # 파일명에서 LotID와 BU/WU를 파싱
    # 패턴 불일치 시 kind=UNKNOWN으로 처리
    m = LOT_PATTERN.match(stem)
    if not m:
        return stem, "UNKNOWN"
    return m.group("lotid"), m.group("kind").upper()


def compute_luminance(rgb) -> float:
    r, g, b = rgb[:3]
    return (0.2126 * r) + (0.7152 * g) + (0.0722 * b)


def compute_red_white_score(rgb) -> float:
    r, g, b = rgb[:3]
    brightness = compute_luminance(rgb) / 255.0
    redness = max(0.0, r - max(g, b)) / 255.0
    whiteness = (min(r, g, b) / 255.0) * brightness
    green_penalty = max(0.0, g - r) / 255.0
    return (redness * 2.2) + (whiteness * 1.4) + (brightness * 0.2) - (green_penalty * 1.0)


def build_safe_sheet_name(base_name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", base_name).strip() or "Sheet"
    candidate = cleaned[:31]
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    idx = 1
    while True:
        suffix = f"_{idx}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        idx += 1


def analyze_bu_grid(
    image_path: Path,
    threshold: int,
    grid_cols: int = BU_GRID_COLS,
    grid_rows: int = BU_GRID_ROWS,
    inner_trim: int = 0,
    crop_box: tuple[int, int, int, int] | None = None,
    analysis_label: str = "",
) -> dict:
    with Image.open(image_path) as img:
        rgb_img = img.convert("RGB")
        width, height = rgb_img.size
        
        # Crop logic
        source_box = (0, 0, width, height)
        if crop_box is not None:
            left, top, right, bottom = crop_box
            if right <= left or bottom <= top:
                return _empty_analysis_dict(grid_rows, grid_cols, inner_trim, (0, 0), crop_box, analysis_label)
            rgb_img = rgb_img.crop((left, top, right, bottom))
            source_box = crop_box
            width, height = rgb_img.size

        if inner_trim > 0:
            left, top, right, bottom = inner_trim, inner_trim, width - inner_trim, height - inner_trim
            if right <= left or bottom <= top:
                return _empty_analysis_dict(grid_rows, grid_cols, inner_trim, (width, height), source_box, analysis_label)
            rgb_img = rgb_img.crop((left, top, right, bottom))
            source_box = (source_box[0] + left, source_box[1] + top, source_box[0] + right, source_box[1] + bottom)
            width, height = rgb_img.size

        # Convert to NumPy for vectorized operations
        arr = np.array(rgb_img, dtype=np.float32)
        # Compute luminance: (0.2126 * r) + (0.7152 * g) + (0.0722 * b)
        # arr shape: (H, W, 3)
        luminance_arr = (0.2126 * arr[:, :, 0]) + (0.7152 * arr[:, :, 1]) + (0.0722 * arr[:, :, 2])
        
        valid_mask = luminance_arr > threshold
        valid_pixels = luminance_arr[valid_mask]

        if valid_pixels.size == 0:
            return _empty_analysis_dict(grid_rows, grid_cols, inner_trim, (width, height), source_box, analysis_label)

        overall_average = np.mean(valid_pixels)
        x_edges = [round(i * width / grid_cols) for i in range(grid_cols + 1)]
        y_edges = [round(i * height / grid_rows) for i in range(grid_rows + 1)]

        cell_averages = [[None for _ in range(grid_cols)] for _ in range(grid_rows)]
        cell_rgb_averages = [[None for _ in range(grid_cols)] for _ in range(grid_rows)]
        cell_deltas = [[None for _ in range(grid_cols)] for _ in range(grid_rows)]
        cell_has_content = [[False for _ in range(grid_cols)] for _ in range(grid_rows)]
        cell_content_ratio = [[0.0 for _ in range(grid_cols)] for _ in range(grid_rows)]
        cell_red_white_scores = [[None for _ in range(grid_cols)] for _ in range(grid_rows)]
        
        valid_cells = 0
        delta_values = []

        for r in range(grid_rows):
            y_start, y_end = y_edges[r], y_edges[r+1]
            for c in range(grid_cols):
                x_start, x_end = x_edges[c], x_edges[c+1]
                
                cell_lum = luminance_arr[y_start:y_end, x_start:x_end]
                cell_rgb = arr[y_start:y_end, x_start:x_end]
                cell_valid_mask = valid_mask[y_start:y_end, x_start:x_end]
                
                # Content ratio based on original pixels in cell
                total_cell_px = cell_lum.size
                if total_cell_px > 0:
                    ratio = np.count_nonzero(cell_valid_mask) / total_cell_px
                    cell_content_ratio[r][c] = ratio
                    cell_has_content[r][c] = ratio >= PRODUCT_CELL_CONTENT_RATIO_MIN
                
                # Use all pixels for basic averages as fallback if needed, but mainly focused on valid ones
                cell_averages[r][c] = np.mean(cell_lum)
                cell_rgb_averages[r][c] = tuple(np.mean(cell_rgb, axis=(0, 1)))
                
                valid_cell_pixels = cell_lum[cell_valid_mask]
                if valid_cell_pixels.size > 0:
                    cell_avg = np.mean(valid_cell_pixels)
                    cell_averages[r][c] = cell_avg
                    delta = overall_average - cell_avg
                    cell_deltas[r][c] = delta
                    delta_values.append(delta)
                    valid_cells += 1
                    
                    # Compute red/white score for the average RGB of valid pixels
                    avg_rgb = np.mean(cell_rgb[cell_valid_mask], axis=0)
                    cell_red_white_scores[r][c] = compute_red_white_score(avg_rgb)

        return {
            "overall_average": float(overall_average),
            "grid_rows": grid_rows,
            "grid_cols": grid_cols,
            "cell_deltas": cell_deltas,
            "cell_averages": cell_averages,
            "cell_rgb_averages": cell_rgb_averages,
            "cell_has_content": cell_has_content,
            "cell_content_ratio": cell_content_ratio,
            "cell_red_white_scores": cell_red_white_scores,
            "valid_cells": valid_cells,
            "valid_pixels": int(valid_pixels.size),
            "min_delta": float(min(delta_values)) if delta_values else None,
            "max_delta": float(max(delta_values)) if delta_values else None,
            "inner_trim": inner_trim,
            "analyzed_size": (width, height),
            "x_edges": x_edges,
            "y_edges": y_edges,
            "source_box": source_box,
            "analysis_label": analysis_label,
        }


def _empty_analysis_dict(grid_rows, grid_cols, inner_trim, size, source_box, label):
    return {
        "overall_average": None,
        "grid_rows": grid_rows,
        "grid_cols": grid_cols,
        "cell_deltas": [[None for _ in range(grid_cols)] for _ in range(grid_rows)],
        "cell_averages": [[None for _ in range(grid_cols)] for _ in range(grid_rows)],
        "cell_rgb_averages": [[None for _ in range(grid_cols)] for _ in range(grid_rows)],
        "cell_has_content": [[False for _ in range(grid_cols)] for _ in range(grid_rows)],
        "cell_content_ratio": [[0.0 for _ in range(grid_cols)] for _ in range(grid_rows)],
        "cell_red_white_scores": [[None for _ in range(grid_cols)] for _ in range(grid_rows)],
        "valid_cells": 0,
        "valid_pixels": 0,
        "min_delta": None,
        "max_delta": None,
        "inner_trim": inner_trim,
        "analyzed_size": size,
        "x_edges": [0 for _ in range(grid_cols + 1)],
        "y_edges": [0 for _ in range(grid_rows + 1)],
        "source_box": source_box,
        "analysis_label": label,
    }


def find_worst_points(analysis: dict, top_n: int = 3) -> list[dict]:
    min_row, max_row, min_col, max_col = get_product_cell_bounds(analysis)
    candidates = []
    for row_idx, row in enumerate(analysis["cell_deltas"], start=1):
        for col_idx, delta in enumerate(row, start=1):
            if delta is None:
                continue
            if not analysis["cell_has_content"][row_idx - 1][col_idx - 1]:
                continue
            if analysis["cell_content_ratio"][row_idx - 1][col_idx - 1] < PRODUCT_CELL_CONTENT_RATIO_MIN:
                continue
            if row_idx < min_row or row_idx > max_row:
                continue
            if col_idx < min_col or col_idx > max_col:
                continue
            score = analysis["cell_red_white_scores"][row_idx - 1][col_idx - 1]
            candidates.append(
                {
                    "row": row_idx,
                    "col": col_idx,
                    "delta": delta,
                    "score": score,
                    "coord": f"({col_idx},{row_idx})",
                }
            )

    candidates.sort(key=lambda item: (-(item["score"] if item["score"] is not None else float("-inf")), item["row"], item["col"]))
    return candidates[:top_n]


def get_product_cell_bounds(analysis: dict) -> tuple[int, int, int, int]:
    min_row = 1
    max_row = analysis["grid_rows"]
    min_col = 1
    max_col = analysis["grid_cols"]

    ratio_map = analysis.get("cell_content_ratio") or []
    valid_rows = []
    valid_cols = []
    for row_idx, row in enumerate(ratio_map, start=1):
        filled_ratio = (sum(1 for value in row if value >= PRODUCT_CELL_CONTENT_RATIO_MIN) / len(row)) if row else 0.0
        if filled_ratio >= PRODUCT_ROW_COVERAGE_MIN:
            valid_rows.append(row_idx)
    for col_idx in range(analysis["grid_cols"]):
        filled_ratio = (
            sum(1 for row_idx in range(len(ratio_map)) if ratio_map[row_idx][col_idx] >= PRODUCT_CELL_CONTENT_RATIO_MIN)
            / len(ratio_map)
        ) if ratio_map else 0.0
        if filled_ratio >= PRODUCT_COL_COVERAGE_MIN:
            valid_cols.append(col_idx + 1)

    if valid_rows:
        min_row = max(min_row, min(valid_rows) - PRODUCT_BOUND_EXPAND_CELLS)
        max_row = min(max_row, max(valid_rows) + PRODUCT_BOUND_EXPAND_CELLS)
    if valid_cols:
        min_col = max(min_col, min(valid_cols) - PRODUCT_BOUND_EXPAND_CELLS)
        max_col = min(max_col, max(valid_cols) + PRODUCT_BOUND_EXPAND_CELLS)

    if min_row > max_row:
        min_row, max_row = 1, analysis["grid_rows"]
    if min_col > max_col:
        min_col, max_col = 1, analysis["grid_cols"]
    return min_row, max_row, min_col, max_col


def get_worst_point_candidate_rect(analysis: dict, width: int, height: int) -> tuple[int, int, int, int]:
    source_box = analysis.get("source_box") or (0, 0, width, height)
    x_edges = analysis.get("x_edges", [])
    y_edges = analysis.get("y_edges", [])
    min_row, max_row, min_col, max_col = get_product_cell_bounds(analysis)
    left_idx = max(0, min_col - 1)
    right_idx = min(len(x_edges) - 1, max_col)
    top_idx = max(0, min_row - 1)
    bottom_idx = min(len(y_edges) - 1, max_row)

    left = source_box[0] + x_edges[left_idx]
    top = source_box[1] + y_edges[top_idx]
    right = source_box[0] + x_edges[right_idx] - 1
    bottom = source_box[1] + y_edges[bottom_idx] - 1
    return (
        max(0, min(left, width - 1)),
        max(0, min(top, height - 1)),
        max(0, min(right, width - 1)),
        max(0, min(bottom, height - 1)),
    )


def get_worst_point_candidate_crop_box(analysis: dict) -> tuple[int, int, int, int]:
    source_box = analysis.get("source_box")
    if source_box is not None:
        width = max(1, source_box[2])
        height = max(1, source_box[3])
    else:
        analyzed_width, analyzed_height = analysis.get("analyzed_size", (1, 1))
        width = max(1, analyzed_width + (analysis.get("inner_trim", 0) * 2))
        height = max(1, analyzed_height + (analysis.get("inner_trim", 0) * 2))
    left, top, right, bottom = get_worst_point_candidate_rect(analysis, width, height)
    return (left, top, right + 1, bottom + 1)


def get_visual_rect(rect: tuple[int, int, int, int], width: int, height: int) -> tuple[int, int, int, int]:
    inset = max(12, int(min(width, height) * 0.02))
    left, top, right, bottom = rect
    return (
        min(right - 1, left + inset),
        min(bottom - 1, top + inset),
        max(left + 1, right - inset),
        max(top + 1, bottom - inset),
    )


def build_worst_point_overlay(image_path: Path, analysis: dict, overlay_path: Path) -> tuple[Path, list[dict]]:
    worst_points = find_worst_points(analysis, top_n=3)
    with Image.open(image_path) as img:
        overlay = img.convert("RGB")
        draw = ImageDraw.Draw(overlay)
        width, height = overlay.size
        source_box = analysis.get("source_box") or (0, 0, width, height)

        # worst point 후보로 인정되는 제품 내부 안전영역만 녹색 사각형으로 표시
        draw.rectangle(get_worst_point_candidate_rect(analysis, width, height), outline=(0, 220, 90), width=3)

        x_edges = analysis.get("x_edges", [])
        y_edges = analysis.get("y_edges", [])
        for rank, point in enumerate(worst_points, start=1):
            col_idx = point["col"] - 1
            row_idx = point["row"] - 1
            center_x = source_box[0] + int((x_edges[col_idx] + x_edges[col_idx + 1]) / 2)
            center_y = source_box[1] + int((y_edges[row_idx] + y_edges[row_idx + 1]) / 2)
            point["pixel_x"] = center_x
            point["pixel_y"] = center_y

            radius = 4
            draw.ellipse(
                (center_x - radius, center_y - radius, center_x + radius, center_y + radius),
                fill=(220, 20, 20),
                outline=(255, 255, 255),
                width=1,
            )
            draw.text((center_x + 10, center_y - 10), f"{rank}:{point['coord']}", fill=(220, 20, 20))

        overlay.save(overlay_path)
    return overlay_path, worst_points


def get_point_center(analysis: dict, point: dict, width: int, height: int) -> tuple[int, int]:
    source_box = analysis.get("source_box") or (0, 0, width, height)
    x_edges = analysis.get("x_edges", [])
    y_edges = analysis.get("y_edges", [])
    col_idx = point["col"] - 1
    row_idx = point["row"] - 1
    center_x = source_box[0] + int((x_edges[col_idx] + x_edges[col_idx + 1]) / 2)
    center_y = source_box[1] + int((y_edges[row_idx] + y_edges[row_idx + 1]) / 2)
    return center_x, center_y


def build_summary_worst_overlay(
    image_path: Path,
    analysis: dict,
    points: list[dict],
    overlay_path: Path,
    label_mode: str = "count_only",
) -> Path:
    with Image.open(image_path) as img:
        overlay = img.convert("RGB")
        draw = ImageDraw.Draw(overlay)
        width, height = overlay.size
        draw.rectangle(get_visual_rect(get_worst_point_candidate_rect(analysis, width, height), width, height), outline=(0, 220, 90), width=3)

        for point in points:
            center_x, center_y = get_point_center(analysis, point, width, height)
            radius = 3
            draw.ellipse(
                (center_x - radius, center_y - radius, center_x + radius, center_y + radius),
                fill=(220, 20, 20),
                outline=(255, 255, 255),
                width=1,
            )
            if label_mode == "count_only":
                label = str(point.get("count", ""))
            else:
                label = point.get("coord", "")
                if point.get("count") is not None:
                    label = f"{label} x{point['count']}"
            draw.text((center_x + 8, center_y - 8), label, fill=(220, 20, 20))

        overlay.save(overlay_path)
    return overlay_path


def build_summary_worst_heatmap(
    image_path: Path,
    analysis: dict,
    points: list[dict],
    overlay_path: Path,
) -> Path:
    with Image.open(image_path) as img:
        width, height = img.size

    worst_rect = get_visual_rect(get_worst_point_candidate_rect(analysis, width, height), width, height)
    records = []
    for point in points:
        center_x, center_y = get_point_center(analysis, point, width, height)
        records.append(
            {
                "X": center_x,
                "Y": center_y,
                "Count": point.get("count", 1),
            }
        )

    if not records:
        Image.new("RGB", (width, height), (0, 0, 0)).save(overlay_path)
        return overlay_path

    df = pd.DataFrame(records)
    plt.style.use("dark_background")
    fig, ax = plt.subplots(figsize=(14, 8), dpi=150)
    ax.set_facecolor("#000000")
    fig.patch.set_facecolor("#000000")
    clip_rect = Rectangle(
        (worst_rect[0], worst_rect[1]),
        worst_rect[2] - worst_rect[0],
        worst_rect[3] - worst_rect[1],
        transform=ax.transData,
    )

    counts = df["Count"].to_numpy(dtype=float)
    count_min = counts.min() if len(counts) else 0.0
    count_max = counts.max() if len(counts) else 1.0
    count_range = max(1e-6, count_max - count_min)
    norm_counts = (counts - count_min) / count_range
    cmap = plt.cm.get_cmap("magma")
    colors = cmap(0.35 + (norm_counts * 0.65))

    # 바깥 glow
    glow_outer = ax.scatter(
        df["X"],
        df["Y"],
        s=500 + (norm_counts * 2200),
        c=colors,
        alpha=0.10,
        marker="o",
        linewidths=0,
    )
    glow_outer.set_clip_path(clip_rect)

    # 중간 glow
    glow_mid = ax.scatter(
        df["X"],
        df["Y"],
        s=180 + (norm_counts * 900),
        c=colors,
        alpha=0.22,
        marker="o",
        linewidths=0,
    )
    glow_mid.set_clip_path(clip_rect)

    # 중심 원형 포인트
    core = ax.scatter(
        df["X"],
        df["Y"],
        s=28 + (norm_counts * 120),
        c=colors,
        alpha=0.95,
        marker="o",
        linewidths=0.4,
        edgecolors="white",
    )
    core.set_clip_path(clip_rect)

    rect = Rectangle(
        (worst_rect[0], worst_rect[1]),
        worst_rect[2] - worst_rect[0],
        worst_rect[3] - worst_rect[1],
        linewidth=1.5,
        edgecolor="#555555",
        facecolor="none",
        linestyle="--",
    )
    ax.add_patch(rect)

    x_margin = width * 0.05
    y_margin = height * 0.05
    ax.set_xlim(-x_margin, width + x_margin)
    ax.set_ylim(height + y_margin, -y_margin)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("DISPLAY BU WORST POINT DENSITY MAP", fontsize=16, fontweight="bold", pad=25, color="white")

    plt.tight_layout()
    plt.savefig(overlay_path, bbox_inches="tight", pad_inches=0.3, facecolor="#000000")
    plt.close(fig)
    return overlay_path


def write_bu_analysis_excel(
    records,
    latest_measurements: dict[str, dict],
    analysis_excel_path: Path,
    threshold: int,
    cancel_check=None,
) -> int:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "요약"
    summary_ws.sheet_view.showGridLines = False
    summary_ws.append(
        [
            "LotID",
            "ModelName",
            "판정",
            "BU data 수치화",
            "Trim5Avg",
            "유효셀수(5px)",
            "최소편차(5px)",
            "최대편차(5px)",
            "Worst1",
            "Worst2",
            "Worst3",
            "분석위치",
            "분석상태",
        ]
    )
    detail_ws = wb.create_sheet("BU_Grid_전체")
    detail_ws.sheet_view.showGridLines = False
    detail_ws["A1"] = "BU Grid Analysis"
    detail_ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    detail_ws["A1"].fill = PatternFill("solid", fgColor="111827")
    detail_ws["A2"] = "크롭 기준"
    detail_ws["B2"] = "검은색 제외 영역의 최소 사각형(BBox)을 찾은 뒤, 사용자 입력 padding을 더해 크롭"
    detail_ws["A3"] = "Grid 기준"
    detail_ws["B3"] = f"크롭된 BU 이미지에서 HSV+Contour로 제품영역을 다시 잡은 뒤 {BU_GRID_COLS} x {BU_GRID_ROWS} 분할"
    detail_ws["A4"] = "편차 계산식"
    detail_ws["B4"] = "편차 = 전체평균밝기 - 셀평균밝기, 밝음=- / 어두움=+"
    detail_ws["A5"] = "비검정 기준"
    detail_ws["B5"] = f"밝기 > threshold({threshold}) 인 픽셀만 사용"
    detail_ws["A6"] = "Grid Data"
    detail_ws["B6"] = "최적화된 제품영역 + 5px 내부 축소 영역"
    detail_ws["A7"] = "Worst Point 기준"
    detail_ws["B7"] = "빨강/흰색 성분 우선 + 제품 content 비율 기준"

    bu_records = [
        rec for rec in records
        if rec.get("kind") == "BU" and rec.get("dst") is not None and str(rec.get("status", "")).startswith("OK")
    ]
    total = len(bu_records)
    print(f"\n[7/8] BU 영역 분석 시작 (대상: {total}개)")

    analysis_count = 0
    detail_start_row = 9
    worst_point_frequency: dict[str, int] = {}
    all_worst_points: list[dict] = []
    summary_overlay_base: Path | None = None
    summary_overlay_analysis: dict | None = None

    for idx, rec in enumerate(sorted(bu_records, key=lambda item: item["lot_id"]), start=1):
        ensure_not_cancelled(cancel_check)
        lot_id = rec["lot_id"]
        measurement = latest_measurements.get(lot_id, {})
        model_name = measurement.get("model_name", "")
        refined_grid_bbox = get_refined_product_bbox(Path(rec["dst"]), margin_px=GRID_REFINED_MARGIN_PX)

        analyses = {
            "trim5": analyze_bu_grid(
                Path(rec["dst"]),
                threshold,
                inner_trim=5,
                crop_box=refined_grid_bbox,
                analysis_label="Refined Product Area + Inner Trim 5px",
            ),
        }
        analysis = analyses["trim5"]
        grid_variants = [
            ("trim5", "Grid Data | Refined Product Area + Inner Trim 5px", "0F766E"),
        ]
        if analysis["overall_average"] is None:
            summary_ws.append(
                [
                    lot_id,
                    model_name,
                    measurement.get("judge", ""),
                    measurement.get("black_uniformity", ""),
                    "",
                    0,
                    "",
                    "",
                    "",
                    "",
                    "",
                    f"row {detail_start_row}",
                    "NO_VALID_PIXEL",
                ]
            )
        else:
            overlay_path = Path(rec["dst"]).with_name(f"{Path(rec['dst']).stem}_worst_overlay{Path(rec['dst']).suffix}")
            overlay_path, worst_points = build_worst_point_overlay(Path(rec["dst"]), analysis, overlay_path)
            for point in worst_points:
                worst_point_frequency[point["coord"]] = worst_point_frequency.get(point["coord"], 0) + 1
                all_worst_points.append(
                    {
                        "coord": point["coord"],
                        "row": point["row"],
                        "col": point["col"],
                        "lot_id": lot_id,
                    }
                )
            if summary_overlay_base is None:
                summary_overlay_base = Path(rec["dst"])
                summary_overlay_analysis = analysis

            detail_ws.merge_cells(start_row=detail_start_row, start_column=1, end_row=detail_start_row, end_column=8)
            detail_ws.cell(row=detail_start_row, column=1, value=f"{lot_id} | {model_name or 'Model N/A'}")
            detail_ws.cell(row=detail_start_row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
            detail_ws.cell(row=detail_start_row, column=1).fill = PatternFill("solid", fgColor="1F2937")

            detail_ws.cell(row=detail_start_row + 1, column=1, value="LotID")
            detail_ws.cell(row=detail_start_row + 1, column=2, value=lot_id)
            detail_ws.cell(row=detail_start_row + 2, column=1, value="ModelName")
            detail_ws.cell(row=detail_start_row + 2, column=2, value=model_name)
            detail_ws.cell(row=detail_start_row + 3, column=1, value="판정")
            detail_ws.cell(row=detail_start_row + 3, column=2, value=measurement.get("judge", ""))
            detail_ws.cell(row=detail_start_row + 4, column=1, value="BU data 수치화")
            detail_ws.cell(row=detail_start_row + 4, column=2, value=measurement.get("black_uniformity", ""))
            detail_ws.cell(row=detail_start_row + 1, column=4, value="전체평균밝기")
            detail_ws.cell(row=detail_start_row + 1, column=5, value=analysis["overall_average"])
            detail_ws.cell(row=detail_start_row + 2, column=4, value="유효셀수")
            detail_ws.cell(row=detail_start_row + 2, column=5, value=analysis["valid_cells"])
            detail_ws.cell(row=detail_start_row + 3, column=4, value="크롭 BBox")
            detail_ws.cell(row=detail_start_row + 3, column=5, value=str(rec.get("bbox", "")))
            detail_ws.cell(row=detail_start_row + 4, column=4, value="설명")
            detail_ws.cell(row=detail_start_row + 4, column=5, value="음수=더 밝음 / 양수=더 어두움")
            detail_ws.cell(row=detail_start_row + 5, column=4, value="Grid Data 기준")
            detail_ws.cell(row=detail_start_row + 5, column=5, value="최적화 제품영역 + 5px 내부 축소")
            detail_ws.cell(row=detail_start_row + 1, column=7, value="Worst1")
            detail_ws.cell(row=detail_start_row + 1, column=8, value=worst_points[0]["coord"] if len(worst_points) >= 1 else "")
            detail_ws.cell(row=detail_start_row + 2, column=7, value="Worst2")
            detail_ws.cell(row=detail_start_row + 2, column=8, value=worst_points[1]["coord"] if len(worst_points) >= 2 else "")
            detail_ws.cell(row=detail_start_row + 3, column=7, value="Worst3")
            detail_ws.cell(row=detail_start_row + 3, column=8, value=worst_points[2]["coord"] if len(worst_points) >= 3 else "")

            if overlay_path.exists():
                bu_img = get_resized_xl_image(overlay_path, 280)
                if bu_img:
                    detail_ws.add_image(bu_img, f"A{detail_start_row + 6}")

            section_top_row = detail_start_row + 6
            grid_start_col = 7
            for trim_index, (analysis_key, block_title, block_color) in enumerate(grid_variants):
                trim_analysis = analyses[analysis_key]
                title_row = section_top_row + trim_index * (BU_GRID_ROWS + 3)
                grid_header_row = title_row + 1

                detail_ws.merge_cells(
                    start_row=title_row,
                    start_column=grid_start_col - 1,
                    end_row=title_row,
                    end_column=grid_start_col + 6,
                )
                detail_ws.cell(row=title_row, column=grid_start_col - 1, value=block_title)
                detail_ws.cell(row=title_row, column=grid_start_col - 1).font = Font(bold=True, color="FFFFFF")
                detail_ws.cell(row=title_row, column=grid_start_col - 1).fill = PatternFill("solid", fgColor=block_color)
                detail_ws.cell(row=title_row, column=grid_start_col + 8, value="Avg")
                detail_ws.cell(row=title_row, column=grid_start_col + 9, value=trim_analysis["overall_average"])
                detail_ws.cell(row=title_row, column=grid_start_col + 10, value="Size")
                detail_ws.cell(row=title_row, column=grid_start_col + 11, value=str(trim_analysis["analyzed_size"]))

                detail_ws.cell(row=grid_header_row, column=grid_start_col - 1, value="Row\\Col")
                for col_idx in range(trim_analysis["grid_cols"]):
                    detail_ws.cell(row=grid_header_row, column=grid_start_col + col_idx, value=col_idx + 1)

                for row_idx in range(trim_analysis["grid_rows"]):
                    detail_ws.cell(row=grid_header_row + 1 + row_idx, column=grid_start_col - 1, value=row_idx + 1)
                    for col_idx in range(trim_analysis["grid_cols"]):
                        delta = trim_analysis["cell_deltas"][row_idx][col_idx]
                        cell = detail_ws.cell(row=grid_header_row + 1 + row_idx, column=grid_start_col + col_idx, value=delta)
                        if delta is not None:
                            cell.number_format = "0.00"

                data_start_col = grid_start_col
                data_end_col = grid_start_col + trim_analysis["grid_cols"] - 1
                data_start_row = grid_header_row + 1
                data_end_row = grid_header_row + trim_analysis["grid_rows"]
                data_range = (
                    f"{get_column_letter(data_start_col)}{data_start_row}:"
                    f"{get_column_letter(data_end_col)}{data_end_row}"
                )
                detail_ws.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=0,
                        mid_color="FFFFFF",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )

            summary_ws.append(
                [
                    lot_id,
                    model_name,
                    measurement.get("judge", ""),
                    measurement.get("black_uniformity", ""),
                    analyses["trim5"]["overall_average"],
                    analyses["trim5"]["valid_cells"],
                    analyses["trim5"]["min_delta"],
                    analyses["trim5"]["max_delta"],
                    worst_points[0]["coord"] if len(worst_points) >= 1 else "",
                    worst_points[1]["coord"] if len(worst_points) >= 2 else "",
                    worst_points[2]["coord"] if len(worst_points) >= 3 else "",
                    f"BU_Grid_전체 row {detail_start_row}",
                    "OK",
                ]
            )
            analysis_count += 1
        detail_start_row += 6 + (len(grid_variants) * (BU_GRID_ROWS + 3)) + 4

        if idx == 1 or idx % 5 == 0 or idx == total:
            print_progress("  BU 분석 진행", idx, total, done=(idx == total))

    summary_ws.freeze_panes = "A2"
    summary_ws["L2"] = "크롭 기준"
    summary_ws["M2"] = "검은색 제외 영역의 최소 사각형(BBox) + padding"
    summary_ws["L3"] = "Grid 기준"
    summary_ws["M3"] = f"{BU_GRID_COLS} x {BU_GRID_ROWS}"
    summary_ws["L4"] = "편차 부호"
    summary_ws["M4"] = "밝음=- / 어두움=+"
    summary_ws["L5"] = "비검정 기준"
    summary_ws["M5"] = f"밝기 > threshold({threshold})"
    summary_ws["L6"] = "Grid Data 기준"
    summary_ws["M6"] = "최적화 제품영역 + 5px 내부 축소"
    summary_ws["L7"] = "Worst Point 기준"
    summary_ws["M7"] = "빨강/흰색 성분 우선 + 제품 content 비율 기준"
    summary_ws["AB2"] = "Worst Point Frequency"
    summary_ws["AB3"] = "Coord"
    summary_ws["AC3"] = "Count"
    for idx, (coord, count) in enumerate(
        sorted(worst_point_frequency.items(), key=lambda item: (-item[1], item[0])),
        start=4,
    ):
        summary_ws.cell(row=idx, column=28, value=coord)
        summary_ws.cell(row=idx, column=29, value=count)

    if summary_overlay_base is not None and summary_overlay_analysis is not None and all_worst_points:
        summary_overlay_path = analysis_excel_path.with_name("bu_worst_points_summary_counts.png")
        summary_heatmap_path = analysis_excel_path.with_name("bu_worst_points_summary_heatmap.png")
        aggregate_points = []
        for coord, count in sorted(worst_point_frequency.items(), key=lambda item: (-item[1], item[0])):
            coord_text = coord.strip("()")
            col_str, row_str = coord_text.split(",")
            aggregate_points.append(
                {
                    "coord": coord,
                    "col": int(col_str),
                    "row": int(row_str),
                    "count": count,
                }
            )
        build_summary_worst_overlay(
            summary_overlay_base,
            summary_overlay_analysis,
            aggregate_points,
            summary_overlay_path,
            label_mode="count_only",
        )
        build_summary_worst_heatmap(summary_overlay_base, summary_overlay_analysis, aggregate_points, summary_heatmap_path)
        summary_ws["L1"] = "Worst Point Count Overlay"
        summary_ws["L1"].font = Font(size=13, bold=True, color="111827")
        overlay_img = XLImage(str(summary_overlay_path))
        if overlay_img.width > 420:
            ratio = 420 / overlay_img.width
            overlay_img.width = int(overlay_img.width * ratio)
            overlay_img.height = int(overlay_img.height * ratio)
        summary_ws.add_image(overlay_img, "L2")
        summary_ws["T1"] = "Worst Point Heatmap"
        summary_ws["T1"].font = Font(size=13, bold=True, color="111827")
        heatmap_img = XLImage(str(summary_heatmap_path))
        if heatmap_img.width > 420:
            ratio = 420 / heatmap_img.width
            heatmap_img.width = int(heatmap_img.width * ratio)
            heatmap_img.height = int(heatmap_img.height * ratio)
        summary_ws.add_image(heatmap_img, "T2")
    for col, width in {
        "A": 24,
        "B": 24,
        "C": 12,
        "D": 12,
        "E": 14,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 24,
        "L": 16,
        "M": 16,
        "T": 16,
        "U": 16,
        "AB": 16,
        "AC": 10,
    }.items():
        summary_ws.column_dimensions[col].width = width

    for col in ("A", "D", "L"):
        summary_ws[f"{col}1"].font = Font(bold=True)
    detail_ws.freeze_panes = "G9"
    for row_idx in range(1, detail_start_row):
        detail_ws.row_dimensions[row_idx].height = DETAIL_ROW_HEIGHT
    for col in ("A", "B", "C", "D", "E", "F"):
        detail_ws.column_dimensions[col].width = 14
    for col in ("P", "Q", "R"):
        detail_ws.column_dimensions[col].width = 14
    for col_idx in range(7, 7 + BU_GRID_COLS):
        detail_ws.column_dimensions[get_column_letter(col_idx)].width = 7

    print("\n[8/8] BU 분석 엑셀 저장")
    wb.save(analysis_excel_path)
    print("  BU 분석 엑셀 저장 완료")
    return analysis_count


def crop_images(input_root: Path, output_root: Path, threshold: int, padding: int, cancel_check=None):
    # merge 결과 폴더를 대상으로 자동 크롭 실행
    if output_root.exists():
        shutil.rmtree(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    image_files = [
        p for p in input_root.rglob("*") if p.is_file() and p.suffix.lower() in ALLOWED_EXTENSIONS
    ]
    total_images = len(image_files)
    print(f"\n[5/7] 이미지 크롭 시작 (대상: {total_images}개)")

    records = []
    for idx, src in enumerate(image_files, start=1):
        ensure_not_cancelled(cancel_check)
        rel = src.relative_to(input_root)
        dst_raw = output_root / rel
        dst = dst_raw
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst = unique_file_path(dst)
        renamed = dst != dst_raw

        lot_id, kind = parse_lot_kind(src.stem)

        try:
            with Image.open(src) as img:
                bbox = find_non_black_bbox(img, threshold=threshold)
                if bbox is None:
                    cropped = img.copy()
                    used_bbox = (0, 0, img.width, img.height)
                    status = "NO_OBJECT_DETECTED"
                else:
                    used_bbox = add_padding(bbox, img.width, img.height, padding)
                    cropped = img.crop(used_bbox)
                    status = "OK"

                if dst.suffix.lower() in (".jpg", ".jpeg") and cropped.mode not in ("RGB", "L"):
                    cropped = cropped.convert("RGB")
                cropped.save(dst)

            records.append(
                {
                    "lot_id": lot_id,
                    "kind": kind,
                    "src": src,
                    "dst": dst,
                    "bbox": used_bbox,
                    "status": status,
                    "renamed_on_save": "TRUE" if renamed else "FALSE",
                }
            )
        except Exception as exc:
            records.append(
                {
                    "lot_id": lot_id,
                    "kind": kind,
                    "src": src,
                    "dst": None,
                    "bbox": None,
                    "status": f"ERROR: {exc}",
                    "renamed_on_save": "FALSE",
                }
            )

        if idx == 1 or idx % 10 == 0 or idx == total_images:
            print_progress("  크롭 진행", idx, total_images, done=(idx == total_images))

    return records


def write_excel(
    records,
    excel_path: Path,
    merge_rows=None,
    latest_measurements=None,
    measurement_rows=None,
    image_width_px: int = 240,
    cancel_check=None,
):
    # 결과 시트: LotID별 BU/WU 이미지 배치 (사용자 입력 컬럼은 공란 유지)
    # 상세 시트: 경로/중복 정보 정리 (이미지 없음)
    wb = Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(["LotID", "판정", "BU data 수치화", "BU Image", "WU data", "WU Image"])

    total = len(records)
    print(f"\n[6/7] 엑셀 작성 시작 (행: {total}개)")

    grouped: dict[str, dict[str, Path | None]] = {}
    for rec in records:
        lot_id = rec["lot_id"]
        kind = rec["kind"]
        grouped.setdefault(lot_id, {"BU": None, "WU": None})
        if kind in ("BU", "WU") and rec["dst"] is not None and grouped[lot_id][kind] is None:
            grouped[lot_id][kind] = rec["dst"]

    row = 2
    for lot_id in sorted(grouped.keys()):
        ensure_not_cancelled(cancel_check)
        ws.cell(row=row, column=1, value=lot_id)
        measurement = (latest_measurements or {}).get(lot_id, {})
        ws.cell(row=row, column=2, value=measurement.get("judge", ""))
        ws.cell(row=row, column=3, value=measurement.get("black_uniformity", ""))
        ws.cell(row=row, column=5, value=measurement.get("white_uniformity", ""))

        bu_path = grouped[lot_id]["BU"]
        wu_path = grouped[lot_id]["WU"]
        max_img_height = 0

        if bu_path is not None and Path(bu_path).exists():
            bu_img = get_resized_xl_image(Path(bu_path), image_width_px)
            if bu_img:
                ws.add_image(bu_img, f"D{row}")
                max_img_height = max(max_img_height, bu_img.height)

        if wu_path is not None and Path(wu_path).exists():
            wu_img = get_resized_xl_image(Path(wu_path), image_width_px)
            if wu_img:
                ws.add_image(wu_img, f"F{row}")
                max_img_height = max(max_img_height, wu_img.height)

        ws.row_dimensions[row].height = max(25, int(max_img_height * 0.75))
        row += 1

    last_result_row = max(2, row - 1)
    ws.conditional_formatting.add(
        f"B2:B{last_result_row}",
        FormulaRule(formula=['B2="OK"'], fill=PatternFill("solid", fgColor="D1FAE5")),
    )
    ws.conditional_formatting.add(
        f"B2:B{last_result_row}",
        FormulaRule(formula=['B2="NG"'], fill=PatternFill("solid", fgColor="FEE2E2")),
    )
    ws.conditional_formatting.add(
        f"C2:C{last_result_row}",
        CellIsRule(operator="lessThan", formula=[str(BU_SPEC_MIN)], fill=PatternFill("solid", fgColor="FEF3C7")),
    )
    ws.conditional_formatting.add(
        f"E2:E{last_result_row}",
        CellIsRule(operator="lessThan", formula=[str(WU_SPEC_MIN)], fill=PatternFill("solid", fgColor="FEF3C7")),
    )

    detail_ws = wb.create_sheet("경로_중복정리")
    detail_ws.append(
        [
            "RecordType",
            "LotID",
            "Kind",
            "Status",
            "DuplicationFlag",
            "PathA",
            "PathB",
            "Etc",
        ]
    )

    for idx, rec in enumerate(records, start=1):
        ensure_not_cancelled(cancel_check)
        bbox_text = "" if rec["bbox"] is None else str(rec["bbox"])
        dst_text = "" if rec["dst"] is None else str(rec["dst"])
        detail_ws.append(
            [
                "CROP",
                rec["lot_id"],
                rec["kind"],
                rec["status"],
                rec.get("renamed_on_save", "FALSE"),
                str(rec["src"]),
                dst_text,
                bbox_text,
            ]
        )
        if idx == 1 or idx % 10 == 0 or idx == total:
            print_progress("  엑셀 진행", idx, total, done=(idx == total))

    if merge_rows:
        for row in merge_rows:
            ensure_not_cancelled(cancel_check)
            detail_ws.append(
                [
                    "MERGE",
                    row["lot_id"],
                    "",
                    "",
                    "TRUE" if row.get("selected_latest_final") == "FALSE" else "FALSE",
                    row.get("folder_path", ""),
                    "",
                    f"created={row.get('created_time','')}, modified={row.get('modified_time','')}",
                ]
            )

    if measurement_rows:
        for row in measurement_rows:
            ensure_not_cancelled(cancel_check)
            detail_ws.append(
                [
                    "MEASURE",
                    row["lot_id"],
                    "",
                    row["judge"],
                    "FALSE" if row.get("selected_latest_final") == "TRUE" else "TRUE",
                    row["source_file"],
                    "",
                    f"time={row['time_raw']}, BU={row['black_uniformity']}, WU={row['white_uniformity']}",
                ]
            )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 36
    detail_ws.column_dimensions["A"].width = 12
    detail_ws.column_dimensions["B"].width = 26
    detail_ws.column_dimensions["C"].width = 10
    detail_ws.column_dimensions["D"].width = 24
    detail_ws.column_dimensions["E"].width = 16
    detail_ws.column_dimensions["F"].width = 60
    detail_ws.column_dimensions["G"].width = 60
    detail_ws.column_dimensions["H"].width = 44
    add_visualization_sheet(wb, latest_measurements or {})
    print("\n메인 엑셀 저장")
    wb.save(excel_path)
    print("  메인 엑셀 저장 완료")


def run_pipeline(integrated_root: Path, data_root: Path, threshold: int, padding: int, cancel_check=None) -> dict:
    # GUI/CLI 공용 실행 함수
    merged_root = integrated_root.parent / f"{integrated_root.name}_LotID_latest_v1"
    cropped_root = integrated_root.parent / f"{integrated_root.name}_LotID_latest_v1_cropped_v1"
    excel_path = cropped_root / "crop_report.xlsx"
    bu_analysis_excel_path = cropped_root / "bu_grid_analysis.xlsx"

    print(f"\n📌 원본 통합 폴더: {integrated_root}")
    print(f"📌 측정 데이터 폴더: {data_root}")
    print(f"📌 정리 폴더(merge): {merged_root}")
    print(f"📌 크롭 폴더: {cropped_root}")
    print(f"📌 엑셀 리포트: {excel_path}")
    print(f"📌 BU 분석 엑셀: {bu_analysis_excel_path}")
    print(f"📌 임계값: {threshold}, 패딩: {padding}")

    ensure_not_cancelled(cancel_check)
    latest_by_lotid, merge_rows = collect_latest_lotid_folders(integrated_root, cancel_check=cancel_check)
    if not latest_by_lotid:
        print("⚠️ LotID 폴더를 찾지 못했어. 폴더 구조를 확인해줘.")
        return

    copy_latest_folders(latest_by_lotid, merged_root, cancel_check=cancel_check)
    merge_report_path = write_merge_report(merge_rows, merged_root)

    latest_measurements, measurement_rows = collect_latest_measurements(data_root, cancel_check=cancel_check)
    crop_records = crop_images(merged_root, cropped_root, threshold, padding, cancel_check=cancel_check)
    write_excel(
        crop_records,
        excel_path,
        merge_rows=merge_rows,
        latest_measurements=latest_measurements,
        measurement_rows=measurement_rows,
        cancel_check=cancel_check,
    )
    bu_analysis_count = write_bu_analysis_excel(
        crop_records,
        latest_measurements,
        bu_analysis_excel_path,
        threshold,
        cancel_check=cancel_check,
    )

    duplicate_count = len(merge_rows) - len(latest_by_lotid)
    ok_count = sum(1 for r in crop_records if r["status"] == "OK")
    nodetect_count = sum(1 for r in crop_records if r["status"] == "NO_OBJECT_DETECTED")
    error_count = sum(1 for r in crop_records if r["status"].startswith("ERROR"))

    print("\n--- 최종 결과 ---")
    print(f"Merge 스캔 LotID 폴더 수: {len(merge_rows)}")
    print(f"Merge 최종 LotID 수: {len(latest_by_lotid)}")
    print(f"Merge 중복 제외 수: {duplicate_count}")
    print(f"Crop 전체 이미지 수: {len(crop_records)}")
    print(f"측정 최신 LotID 수: {len(latest_measurements)}")
    print(f"Crop 성공: {ok_count}")
    print(f"Crop 객체 미검출(원본 유지): {nodetect_count}")
    print(f"Crop 오류: {error_count}")
    print(f"Merge 리포트: {merge_report_path}")
    print(f"Crop 엑셀: {excel_path}")
    print(f"BU 분석 엑셀: {bu_analysis_excel_path}")
    print(f"BU 분석 완료 LotID 수: {bu_analysis_count}")
    print("완료!")

    return {
        "merged_root": merged_root,
        "cropped_root": cropped_root,
        "excel_path": excel_path,
        "bu_analysis_excel_path": bu_analysis_excel_path,
        "merge_report_path": merge_report_path,
        "merge_rows": len(merge_rows),
        "latest_lotids": len(latest_by_lotid),
        "latest_measurements": len(latest_measurements),
        "crop_records": len(crop_records),
        "crop_ok": ok_count,
        "crop_nodetect": nodetect_count,
        "crop_error": error_count,
        "bu_analysis_count": bu_analysis_count,
    }


def main():
    # 원클릭 실행 순서:
    # 1) 사용자 입력 수집
    # 2) LotID 최신 폴더 취합(merge)
    # 3) 측정 데이터 최신값 집계
    # 4) merge 결과를 자동 크롭
    # 5) 엑셀 리포트 생성
    print("\n--- BU Organize One Click v1 ---")
    integrated_root = ask_path("1) 이미지 통합 폴더 경로: ")
    if not integrated_root.exists() or not integrated_root.is_dir():
        print(f"🚨 폴더를 찾을 수 없어: {integrated_root}")
        return

    data_root = ask_path("2) 측정 데이터 상위 폴더 경로: ")
    if not data_root.exists() or not data_root.is_dir():
        print(f"🚨 측정 데이터 폴더를 찾을 수 없어: {data_root}")
        return

    threshold = ask_int("3) 비검정 판정 임계값(0~255)", 12)
    padding = ask_int("4) 크롭 패딩(px)", 8)
    run_pipeline(integrated_root, data_root, threshold, padding)


if __name__ == "__main__":
    main()
